#!/usr/bin/python3

# Author: Connor McLeod
# Contact: con.mcleod92@gmail.com
# Source code: https://github.com/con-mcleod/MonthlyPerf_Report
# Latest Update: 10 August 2018

import sys, sqlite3, os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, Font, PatternFill, Border, Side

##############################
#                            #
# Database handling          #
#                            #
##############################

def dbselect(cxn, query, payload):
	"""
	Function to select data from an sqlite3 table
	:param cxn: connection to the sqlite3 database
	:param query: the query to be run
	:param payload: the payload for any query parameters
	:return results: the results of the search
	"""
	cursor = cxn.cursor()
	if not payload:
		rows = cursor.execute(query)
	else:
		rows = cursor.execute(query,payload)
	results = []
	for row in rows:
		results.append(row)
	cursor.close()
	return results

##############################
#                            #
# Helper functions           #
#                            #
##############################

def get_all_SMIs(cxn):
	"""
	Function to grab all SMIs from the Encompass reports
	:param cxn: connection to sqlite3 database
	:return all_SMIs: list of all SMIs
	"""
	query = "SELECT distinct(SMI) from DAILY_GEN"
	payload = None
	all_SMIs = dbselect(cxn, query, payload)
	return all_SMIs


def get_all_months(cxn):
	"""
	Function to return all months included in the Encompass reports
	:param cxn: connection to sqlite3 database
	:return all_dates: list of all months in report [mm, yy]
	"""
	query = """SELECT obs_month, obs_year from DAILY_GEN 
			group by obs_month, obs_year order by obs_year, obs_month"""
	payload = None
	all_dates = dbselect(cxn, query, payload)
	return all_dates


def get_last_date(cxn):
	"""
	Function to return the date of the last data entry from Encompass reports
	:param cxn: connection to sqlite3 database
	:return last_date: last date of Encompass reports
	"""
	query = """SELECT obs_day, obs_month, obs_year from DAILY_GEN
			group by obs_day, obs_month, obs_year order by obs_year, obs_month, obs_day"""
	payload = None
	all_dates = dbselect(cxn, query, payload)
	last_date = all_dates[-1]
	return last_date


def get_SMI_details(cxn, SMI):
	"""
	Function to return an SMI's Salesforce details
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:return result: list of SMI's Salesforce details
	"""
	query = """SELECT ref_no, state, installer, PVsize, export_control,
		panel_brand, site_type, site_status, supply_date, tariff from SMI_DETAILS where SMI=?"""
	payload = (SMI)
	result = dbselect(cxn, query, payload)
	if not result:
		for i in range(0,22):
			result.append((''))
	return result


def get_SMI_forecast(cxn, SMI):
	"""
	Function to return an SMI's forecast data
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:return result: list of forecast values
	"""
	query = "SELECT val from FORECAST where SMI=?"
	payload = (SMI)
	result = dbselect(cxn, query, payload)
	return result



def get_SMI_adj_forecast(cxn, SMI):
	"""
	Function to return an SMI's adjusted forecast data
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:return result: list of adjusted forecast values
	"""
	query = "SELECT adj_val from ADJ_FORECAST where SMI=?"
	payload = (SMI)
	result = dbselect(cxn, query, payload)
	return result


def get_SMI_generation(cxn, SMI):
	"""
	Function to return an SMI's adjusted monthly generation
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:return result: list of SMI's adjusted monthly generation
	"""
	query = "SELECT val from MONTH_GEN where SMI=? order by year, month"
	payload = (SMI)
	result = dbselect(cxn, query, payload)
	if not result:
		for i in range(0,len(dates)):
			result.append((0,))
	return result


def get_perf(cxn, SMI, period):
	"""
	Function to return an SMI's adjusted monthly generation
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:param period: the period of interest (annual, quarter, month, prev month)
	:return fc: list of forecasts for each month in period
	:return gen: list of generation values for each month in period
	:return perf: list of generation/forecast as a % for each month in period
	"""
	fc = 0
	gen = 0
	query = "SELECT adj_val from ADJ_FORECAST where SMI=? order by year, month"
	payload = (SMI)
	fc_vals = dbselect(cxn, query, payload)
	if (period == "Annual"):
		fc_vals = fc_vals[-12:]
	elif (period == "Quarter"):
		fc_vals = fc_vals[-3:]
	elif (period == "Month"):
		fc_vals = fc_vals[-1:]

	if fc_vals:
		if (period == "Prev"):
			fc_vals = fc_vals[-2]

	query = "SELECT val from MONTH_GEN where SMI=? order by year, month"
	payload = (SMI)
	gen_vals = dbselect(cxn, query, payload)[-12:]
	if (period == "Annual"):
		gen_vals = gen_vals[-12:]
	elif (period == "Quarter"):
		gen_vals = gen_vals[-3:]
	elif (period == "Month"):
		gen_vals = gen_vals[-1:]
	
	if gen_vals:
		if (period == "Prev"):
			gen_vals = gen_vals[-2]

	if period == "Prev":
		for fc_val in fc_vals:
			fc += fc_val
		for gen_val in gen_vals:
			gen += gen_val
		if (fc != 0):
			perf = gen/fc
		else:
			perf = 0
	else:
		for fc_val in fc_vals:
			fc += fc_val[0]
		for gen_val in gen_vals:
			gen += gen_val[0]
		if (fc != 0):
			perf = gen/fc
		else:
			perf = 0
	return fc, gen, perf


def get_off_days(cxn, SMI, dates):
	"""
	Function to return the number of days a site had zero generation in a given month
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:param dates: all dates given from Encompass files
	:return off_days: number of days of zero generation in the current month for an SMI
	"""
	curr_date = dates[-1]
	curr_month = curr_date[0]
	curr_year = curr_date[1]
	off_days = 0
	query = """SELECT value from DAILY_GEN where SMI=? and obs_month=? 
				and obs_year=?"""
	payload = (SMI[0], curr_month, curr_year)
	result = dbselect(cxn, query, payload)
	if result:
		for val in result:
			if not val[0]:
				off_days += 1
			elif val[0] < 0.1:
				off_days += 1
	return off_days


##############################
#                            #
# GENERATE OUTPUT            #
#                            #
##############################

if __name__ == '__main__':

	# terminate program if not executed correctly
	if (len(sys.argv) != 1):
		print ("Usage: python3 genMonthlyReport.py")
		exit(1)

	# connect to the sqlite3 database
	DATABASE = "dataset.db"
	cxn = sqlite3.connect(DATABASE)

	# name the output file using YY.MM.DD.xlsx format
	last_date = get_last_date(cxn)
	output = str(last_date[2])+"."+str(last_date[1])+"."+str(last_date[0]) + ".xlsx"

	# open the output file because it should already exist from genAllSites.py
	wb = load_workbook(output)
	if 'Perf Report' in wb.sheetnames:
		ws = wb.get_sheet_by_name('Perf Report')
	else:
		ws = wb.create_sheet('Perf Report')

	# openpyxl cell formatting
	redFill = PatternFill(start_color='FA5858', end_color='FA5858', fill_type='solid')
	greenFill = PatternFill(start_color='9Afe2e', end_color='9Afe2e', fill_type='solid')
	leftBorder = Border(left=Side(style='thin'))
	rightBorder = Border(right=Side(style='thin'))

	# grab column and row data types
	SMIs = get_all_SMIs(cxn)
	dates = get_all_months(cxn)

	# create the headings
	ws_headings = ["SMI","Ref No","State","Installer","System Size","Export Control",
					"Panel Make","System Type","PPA Status","Supply Date","Tariff","Jan FC","Feb FC", "Mar FC",
					"Apr FC","May FC","Jun FC","Jul FC","Aug FC","Sep FC","Oct FC",
					"Nov FC","Dec FC"]
	for date in dates:
		date = "adj_fc(" + str(date).strip('()') + ")"
		ws_headings.append(date)
	for date in dates:
		date = "gen(" + str(date).strip('()') + ")"
		ws_headings.append(date)
	ws_headings.extend(["Annual FC","Annual Gen","Annual Perf","Quarter FC","Quarter Gen",
						"Quater Perf","Month FC","Month Gen","Month Perf","Prev FC",
						"Prev Gen","Prev Perf","Outage Days","Annual FC $","Annual Gen $","Shortfall $",
						"Quarter FC $","Quarter Gen $","Shortfall $","CurrMonth FC $","CurrMonth Gen $",
						"Shortfall $","PrevMonth FC $","PrevMonth Gen $","Shortfall $"])

	# loop through each SMI and add the data
	row_count = 1
	for SMI in SMIs:
		print ("Formatting SMI: " + SMI[0])
		col_count = 0

		# print headings
		if row_count == 1:
			for heading in ws_headings:
				ws.cell(row=row_count, column=col_count+1).value = heading
				col_count += 1
		col_count = 1
		
		ws.cell(row=row_count+1, column=1).value = SMI[0]
		
		# add Salesforce details
		details = get_SMI_details(cxn, SMI)
		if details:
			if details[0]:
				for detail in details[0]:
					ws.cell(row=row_count+1, column=col_count+1).value = detail
					col_count += 1
			else:
				for i in range(0,len(details)):
					ws.cell(row=row_count+1, column=col_count+1).value = ''
					col_count += 1
		
		# add Salesforce forecasts
		forecasts = get_SMI_forecast(cxn, SMI)
		ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
		for forecast in forecasts:
			ws.cell(row=row_count+1, column=col_count+1).value = forecast[0]
			col_count += 1

		# if no supply date, skip SMI from performance metrics and highlight red
		if not details[0][8]:
			print ("No supply date for: " + SMI[0] + " - skipping")
			ws.cell(row=row_count+1, column=1).fill = redFill
			row_count += 1
			continue

		# add adjusted forecast based on supply date
		adj_forecasts = get_SMI_adj_forecast(cxn, SMI)
		ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
		for adj_forecast in adj_forecasts:
			ws.cell(row=row_count+1, column=col_count+1).value = adj_forecast[0]
			col_count += 1
		
		# add monthly generation from encompass
		month_gen = get_SMI_generation(cxn, SMI)
		ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
		for gen in month_gen:
			ws.cell(row=row_count+1, column=col_count+1).value = gen[0]
			col_count += 1
		
		# add annual performance columns
		i = 0
		annual_perf = get_perf(cxn, SMI, "Annual")
		for val in annual_perf:
			ws.cell(row=row_count+1, column=col_count+1).value = val
			if i == 0:
				ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
			if i == 2:
				ws.cell(row=row_count+1, column=col_count+1).number_format = '0.00%'
				if val < .9:
					ws.cell(row=row_count+1, column=col_count+1).fill = redFill
				elif val > 1.2:
					ws.cell(row=row_count+1, column=col_count+1).fill = greenFill
			col_count += 1
			i += 1

		# add quarter performance columns
		i = 0
		quarter_perf = get_perf(cxn, SMI, "Quarter")
		for val in quarter_perf:
			ws.cell(row=row_count+1, column=col_count+1).value = val
			if i == 0:
				ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
			if i == 2:
				ws.cell(row=row_count+1, column=col_count+1).number_format = '0.00%'
				if val < .9:
					ws.cell(row=row_count+1, column=col_count+1).fill = redFill
				elif val > 1.2:
					ws.cell(row=row_count+1, column=col_count+1).fill = greenFill
			col_count += 1
			i += 1

		# add monthly performance columns
		i = 0
		month_perf = get_perf(cxn, SMI, "Month")
		for val in month_perf:
			ws.cell(row=row_count+1, column=col_count+1).value = val
			if i == 0:
				ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
			if i == 2:
				ws.cell(row=row_count+1, column=col_count+1).number_format = '0.00%'
				if val < .9:
					ws.cell(row=row_count+1, column=col_count+1).fill = redFill
				elif val > 1.2:
					ws.cell(row=row_count+1, column=col_count+1).fill = greenFill
			col_count += 1
			i += 1

		# add last months performance columns
		i = 0
		last_month_perf = get_perf(cxn, SMI, "Prev")
		for val in last_month_perf:
			ws.cell(row=row_count+1, column=col_count+1).value = val
			if i == 0:
				ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
			if i == 2:
				ws.cell(row=row_count+1, column=col_count+1).number_format = '0.00%'
				if val < .9:
					ws.cell(row=row_count+1, column=col_count+1).fill = redFill
				elif val > 1.2:
					ws.cell(row=row_count+1, column=col_count+1).fill = greenFill
				ws.cell(row=row_count+1, column=col_count+1).border = rightBorder
			col_count += 1
			i += 1
		
		# add number of off days
		off_days = get_off_days(cxn, SMI, dates)
		ws.cell(row=row_count+1, column=col_count+1).border = rightBorder
		ws.cell(row=row_count+1, column=col_count+1).value = off_days
		if (off_days):
			if off_days > 0:
				ws.cell(row=row_count+1, column=col_count+1).fill = redFill

		col_count += 1

		# add tariff
		tariff = None
		if (details[0]):
			if details[0][9]:
				tariff = float(details[0][9])
		
		# add revenue impact if tariff exists
		if tariff:
			revenues = []
			for val in annual_perf[:2]:
				revenue = (val*tariff)/100
				ws.cell(row=row_count+1, column=col_count+1).value = revenue
				revenues.append(revenue)
				col_count += 1
			ws.cell(row=row_count+1, column=col_count+1).value = (revenues[1]-revenues[0])
			ws.cell(row=row_count+1, column=col_count+1).border = rightBorder
			col_count += 1

			revenues = []
			for val in quarter_perf[:2]:
				revenue = (val*tariff)/100
				ws.cell(row=row_count+1, column=col_count+1).value = revenue
				revenues.append(revenue)
				col_count += 1
			ws.cell(row=row_count+1, column=col_count+1).value = (revenues[1]-revenues[0])
			ws.cell(row=row_count+1, column=col_count+1).border = rightBorder
			col_count += 1

			revenues = []
			for val in month_perf[:2]:
				revenue = (val*tariff)/100
				ws.cell(row=row_count+1, column=col_count+1).value = revenue
				revenues.append(revenue)
				col_count += 1
			ws.cell(row=row_count+1, column=col_count+1).value = (revenues[1]-revenues[0])
			ws.cell(row=row_count+1, column=col_count+1).border = rightBorder
			col_count += 1

			revenues = []
			for val in last_month_perf[:2]:
				revenue = (val*tariff)/100
				ws.cell(row=row_count+1, column=col_count+1).value = revenue
				revenues.append(revenue)
				col_count += 1
			ws.cell(row=row_count+1, column=col_count+1).value = (revenues[1]-revenues[0])
			ws.cell(row=row_count+1, column=col_count+1).border = rightBorder

		row_count += 1

	# save the excel file
	wb.save(output)

	print ("Complete!")