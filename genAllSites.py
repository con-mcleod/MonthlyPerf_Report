#!/usr/bin/python3

# Author: Connor McLeod
# Contact: con.mcleod92@gmail.com
# Source code: https://github.com/con-mcleod/MonthlyPerf_Report
# Latest Update: 10 August 2018

import sys, sqlite3, os
from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill, Border, Side

##############################
#                            #
# Database handling          #
#                            #
##############################

def dbselect(cxn, query, payload):
	"""
	Function to select data from an sqlite3 table
	:param cxn: connection to the sqlite3 database
	:param query: the query to be run
	:param payload: the payload for any query parameters
	:return results: the results of the search
	"""
	cursor = cxn.cursor()
	if not payload:
		rows = cursor.execute(query)
	else:
		rows = cursor.execute(query,payload)
	results = []
	for row in rows:
		results.append(row)
	cursor.close()
	return results

##############################
#                            #
# Helper functions           #
#                            #
##############################

def get_all_SMIs(cxn):
	"""
	Function to grab all SMIs from the Encompass reports
	:param cxn: connection to sqlite3 database
	:return all_SMIs: list of all SMIs
	"""
	query = "SELECT distinct(SMI) from DAILY_GEN"
	payload = None
	all_SMIs = dbselect(cxn, query, payload)
	return all_SMIs


def get_all_months(cxn):
	"""
	Function to return all months included in the Encompass reports
	:param cxn: connection to sqlite3 database
	:return all_dates: list of all months in report [mm, yy]
	"""
	query = """SELECT obs_month, obs_year from DAILY_GEN 
			group by obs_month, obs_year order by obs_year, obs_month"""
	payload = None
	all_dates = dbselect(cxn, query, payload)
	return all_dates


def get_last_date(cxn):
	"""
	Function to return the date of the last data entry from Encompass reports
	:param cxn: connection to sqlite3 database
	:return last_date: last date of Encompass reports
	"""
	query = """SELECT obs_day, obs_month, obs_year from DAILY_GEN
			group by obs_day, obs_month, obs_year order by obs_year, obs_month, obs_day"""
	payload = None
	all_dates = dbselect(cxn, query, payload)
	last_date = all_dates[-1]
	return last_date


def get_month_gen(cxn, SMI, date):
	"""
	Function to return the monthly generation for an SMI in a given month
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:param date: the month and year of interest
	:return result: the monthly generation
	"""
	month = date[0]
	year = date[1]
	query = "SELECT val from MONTH_GEN where SMI=? and month=? and year=?"
	payload = (SMI, month, year)
	gen = dbselect(cxn, query, payload)
	return gen


def get_off_days(cxn, SMI, dates):
	"""
	Function to return the number of days a site had zero generation in a given month
	:param cxn: connection to sqlite3 database
	:param SMI: the SMI of interest
	:param dates: all dates given from Encompass files
	:return off_days: number of days of zero generation in the current month for an SMI
	"""
	curr_date = dates[-1]
	curr_month = curr_date[0]
	curr_year = curr_date[1]
	off_days = 0
	query = """SELECT value from DAILY_GEN where SMI=? and obs_month=? 
				and obs_year=?"""
	payload = (SMI[0], curr_month, curr_year)
	result = dbselect(cxn, query, payload)
	if result:
		for val in result:
			if not val[0]:
				off_days += 1
			elif val[0] < 0.1:
				off_days += 1
	return off_days

##############################
#                            #
# GENERATE OUTPUT            #
#                            #
##############################

if __name__ == '__main__':

	# terminate program if not executed correctly
	if (len(sys.argv) != 1):
		print ("Usage: python3 genAllSites.py")
		exit(1)

	# connect to the database and create the tables
	DATABASE = "dataset.db"
	cxn = sqlite3.connect(DATABASE)

	# name the output file using YY.MM.DD.xlsx format
	last_date = get_last_date(cxn)
	output = str(last_date[2])+"."+str(last_date[1])+"."+str(last_date[0]) + ".xlsx"

	if (os.path.exists(output)):
		os.remove(output)

	# openpyxl commands to create excel workbook and sheets
	wb = Workbook()
	ws = wb.active
	ws.title = "All_sites"

	# openpyxl format styles
	leftBorder = Border(left=Side(style='thin'))
	rightBorder = Border(right=Side(style='thin'))

	dates = get_all_months(cxn)
	SMIs = get_all_SMIs(cxn)

	# for each SMI format and store the data in the sheet
	row_count = 1
	for SMI in SMIs:
		col_count = 1
		ws.cell(row=1, column=1).value = "SMI"
		ws.cell(row=row_count+1, column=1).value = SMI[0]
		ws.cell(row=row_count+1, column=1).border = rightBorder

		print("Formatting SMI: " + SMI[0])

		for date in dates:
			ws.cell(row=1, column=col_count+1).value = str(date[0]) + "," + str(date[1])
			month_gen = get_month_gen(cxn, SMI[0], date)
			if month_gen:
				ws.cell(row=row_count+1, column=col_count+1).value = month_gen[0][0]
			col_count += 1
		ws.cell(row=1, column=col_count+1).value = "Outage Days"
		off_days = get_off_days(cxn, SMI, dates)
		ws.cell(row=row_count+1, column=col_count+1).border = leftBorder
		ws.cell(row=row_count+1, column=col_count+1).value = off_days
		row_count += 1

	wb.save(output)	

	print ("Complete!")